import requests
import re
import copy
import openpyxl
from bs4 import BeautifulSoup
from tqdm import tqdm

wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.active
sheet.title = 'OSU Clubs'
num_pages = 118
column_names = []

for i in range(23):
    column_names.append(sheet.cell(row=1, column=i+1).value)

sample_dict = dict((col_name,'Not Listed') for col_name in column_names)

def scrape_data():
    row_counter = 1

    for i in tqdm(range(num_pages)):
        page = requests.get(f'https://activities.osu.edu/involvement/student_organizations/find_a_student_org?page={i}&v=card&c=all')

        soup = BeautifulSoup(page.content, 'html.parser')

        cards = soup.find_all("a", {"id": re.compile("ctl00_ContentBody_pageFormControl_rpt_cards_ctl\d\d_OrgCard_HypCard")})

        for c_index, card in enumerate(cards):

            url = f'https://activities.osu.edu{card["href"]}'

            org = requests.get(url)

            soup_org = BeautifulSoup(org.content, 'html.parser')
        

            data = soup_org.find_all("tr")

            data_dict = copy.deepcopy(sample_dict)

            title = soup_org.find('h4').text

            data_dict["Organization Name"] = title

            for row in data:
                th = row.find("th").text.strip()[:-1]
                td = row.find("td").text.strip()
                data_dict[th] = td
            row_counter+=1

            for i, v in enumerate(data_dict.values()):
                cell = sheet.cell(row=row_counter, column=i+1)
                cell.value = v

    wb.save('ready_data.xlsx')

def main():
    scrape_data()

if __name__ == '__main__':
    main()
